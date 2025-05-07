import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel(data):
    # Construct DataFrame
    rows = []
    for blog in data:
        rows.append({
            "Website": blog["website"],
            "Title": blog["title"],
            "Content": blog["content"],
            "Link": blog["link"],
            "Published": blog["metadata"].get("published", "N/A"),
            "VADER Pos": blog["analysis"]["vader"]["pos"],
            "VADER Neu": blog["analysis"]["vader"]["neu"],
            "VADER Neg": blog["analysis"]["vader"]["neg"],
            "Empath Positive": blog["analysis"]["empath"].get("positive_emotion", 0),
            "Empath Negative": blog["analysis"]["empath"].get("negative_emotion", 0),
            "LLM Summary": blog["analysis"]["llm"]
        })

    df = pd.DataFrame(rows)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Blog Analysis")
        workbook = writer.book
        worksheet = writer.sheets["Blog Analysis"]

        # Styling
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="DDEBF7")
        alignment = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_num, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = max(15, len(col_name) + 5)
            cell = worksheet[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

    output.seek(0)
    return output
